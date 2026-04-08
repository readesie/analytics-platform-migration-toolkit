from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Timeline Analysis"

ACCENT  = "1B4F72"
ACCENT2 = "2471A3"
WHITE   = "FFFFFF"
DARK    = "1A1A1A"
MID     = "444444"

GREEN_BG  = "E8F5E9"
GREEN_FG  = "1B5E20"
AMBER_BG  = "FFF8E1"
AMBER_FG  = "E65100"
RED_BG    = "FFEBEE"
RED_FG    = "B71C1C"
BLUE_BG   = "E3F2FD"
BLUE_FG   = "0D47A1"
GRAY_BG   = "F5F5F5"

def fill(h): return PatternFill("solid", fgColor=h)
def s(): return Side(style="thin", color="CCCCCC")
def b(): return Border(left=s(), right=s(), top=s(), bottom=s())
def ms(): return Side(style="medium", color=ACCENT)
def mb(): return Border(left=ms(), right=ms(), top=ms(), bottom=ms())

ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

col_w = {"A":6,"B":28,"C":12,"D":12,"E":38,"F":18}
for col, w in col_w.items():
    ws.column_dimensions[col].width = w

# ── TITLE ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
ws["A1"] = "Northstar Analytics Group MIGRATION  —  REALISTIC TIMELINE ANALYSIS"
ws["A1"].font      = Font(name="Arial", bold=True, color=WHITE, size=14)
ws["A1"].fill      = fill(ACCENT)
ws["A1"].alignment = ctr
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:F2")
ws["A2"] = "6-Month Contract Starting ~April 2026  |  Realistic Range: 6–10 Months  |  Key threats identified below"
ws["A2"].font      = Font(name="Arial", color=WHITE, size=11, italic=True)
ws["A2"].fill      = fill(ACCENT2)
ws["A2"].alignment = ctr
ws.row_dimensions[2].height = 22

ws.row_dimensions[3].height = 8

# ── SECTION: OPTIMISTIC BASELINE ──────────────────────────────────────
ws.merge_cells("A4:F4")
ws["A4"] = "BASELINE — OPTIMISTIC 6-MONTH SCHEDULE  (everything goes right)"
ws["A4"].font      = Font(name="Arial", bold=True, color=WHITE, size=12)
ws["A4"].fill      = fill(GREEN_FG)
ws["A4"].alignment = lft
ws.row_dimensions[4].height = 24

headers = ["#", "Phase", "Start", "End", "Key Assumptions for This to Hold", "Risk if Wrong"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=i, value=h)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill      = fill(ACCENT)
    c.alignment = ctr
    c.border    = b()
ws.row_dimensions[5].height = 28

baseline = [
    (1, "Pre-Start & Access",        "Wk 1",    "Wk 1",
     "Contract signed day 1; access to all environments granted within 48 hours",
     "Access delays alone can cost 2–3 weeks"),
    (2, "Discovery & Dependency Audit", "Wk 1",  "Wk 3",
     "Evan and SAS Admin available; dependency tracking gaps manageable",
     "Fragmented tracking (per Evan) could extend this to 6+ weeks"),
    (3, "Planning & OS Version Confirm", "Wk 3", "Wk 5",
     "Target OS confirmed quickly; IT procurement already in motion",
     "Licensing/procurement is an IT/finance decision — can take weeks"),
    (4, "OS Upgrade (all platforms, all tiers)", "Wk 5", "Wk 10",
     "IT provisions servers fast; SAS/Spotfire/PBI compatible with target OS",
     "Application compatibility testing alone can add 2–4 weeks"),
    (5, "Separation Build — Dev",     "Wk 9",   "Wk 14",
     "New domain infrastructure ready; DB connection scope confirmed",
     "New IT domain setup is an IT project in itself"),
    (6, "Sandbox + UAT",              "Wk 13",  "Wk 18",
     "~140 users available for UAT; sign-off obtained promptly",
     "UAT scheduling with business users is notoriously slippery"),
    (7, "Prod Cutover + Validation",  "Wk 18",  "Wk 22",
     "No critical issues found in Sandbox; go/no-go is clean",
     "One major issue in UAT can push cutover 4–6 weeks"),
    (8, "Stabilization & Closeout",   "Wk 22",  "Wk 26",
     "Post-cutover stable within 72 hours; documentation completed",
     "Undiscovered dependencies surface post-cutover — common"),
]

for idx, (num, phase, start, end, assumption, risk) in enumerate(baseline, 6):
    row_fill = GREEN_BG if idx % 2 == 0 else WHITE
    for ci, val in enumerate([num, phase, start, end, assumption, risk], 1):
        c = ws.cell(row=idx, column=ci, value=val)
        c.fill      = fill(row_fill)
        c.border    = b()
        c.alignment = ctr if ci in (1,2,3,4) else lft
        if ci == 1:
            c.font = Font(name="Arial", bold=True, size=11, color=ACCENT)
        elif ci == 2:
            c.font = Font(name="Arial", bold=True, size=11, color=DARK)
        elif ci == 6:
            c.font = Font(name="Arial", size=10, color=RED_FG, italic=True)
        else:
            c.font = Font(name="Arial", size=10, color=MID)
    ws.row_dimensions[idx].height = 42

ws.row_dimensions[14].height = 12

# ── SECTION: THREATS ──────────────────────────────────────────────────
ws.merge_cells("A15:F15")
ws["A15"] = "THREATS TO THE 6-MONTH TIMELINE  —  Ranked by Likelihood"
ws["A15"].font      = Font(name="Arial", bold=True, color=WHITE, size=12)
ws["A15"].fill      = fill(RED_FG)
ws["A15"].alignment = lft
ws.row_dimensions[15].height = 24

threat_headers = ["Risk", "Threat", "Likelihood", "Weeks at Risk", "Impact", "Mitigation"]
for i, h in enumerate(threat_headers, 1):
    c = ws.cell(row=16, column=i, value=h)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill      = fill(ACCENT)
    c.alignment = ctr
    c.border    = b()
ws.row_dimensions[16].height = 28

threats = [
    ("T1", "Fragmented dependency tracking — no single source of truth (per Evan)",
     "🔴 HIGH", "+2 to +6 wks",
     "Discovery takes much longer than planned; migration built on incomplete map",
     "Phase 1 #9 — audit ALL sources before touching anything"),
    ("T2", "IT provisioning queue for new servers / new domain infrastructure",
     "🔴 HIGH", "+2 to +4 wks",
     "OS upgrade and separation build cannot start until IT provisions environments",
     "Raise provisioning request on Day 1; track daily"),
    ("T3", "OS / target version not confirmed; procurement / licensing delay",
     "🟠 MED-HIGH", "+1 to +4 wks",
     "Planning cannot lock down until version is confirmed; may affect hardware compat",
     "Get written version confirmation in Week 1 kick-off"),
    ("T4", "Application compatibility: SAS, Spotfire, or PBI Gateway not certified on target OS",
     "🟠 MED-HIGH", "+2 to +6 wks",
     "Could require vendor support calls, patches, or version upgrades of the apps themselves",
     "Verify all three platform vendors' OS support matrix before upgrade begins"),
    ("T5", "UAT scheduling delays — ~140 business users hard to coordinate",
     "🟠 MEDIUM", "+1 to +4 wks",
     "Sandbox sign-off gate blocked; Prod cutover window slips",
     "Book UAT window early; get Priya to own user scheduling"),
    ("T6", "SAS admin knowledge gaps — Evan inherited role ~2.5 yrs ago, no formal handoff",
     "🟠 MEDIUM", "+1 to +3 wks",
     "Undocumented SAS configs surface during migration; rework required",
     "Treat Discovery Phase as mandatory — do not skip or rush"),
    ("T7", "Post-cutover undiscovered DB dependencies surface in Prod",
     "🟡 MED-LOW", "+1 to +4 wks",
     "~140 users reporting issues; emergency stabilization period",
     "Consolidated dependency map (#15 in checklist) is the prevention"),
    ("T8", "Contract start delayed (hiring decision, onboarding, security clearance)",
     "🟡 MED-LOW", "+1 to +3 wks",
     "Entire timeline shifts right; 6-month window may not accommodate full scope",
     "Discuss with Palmer Group; push for fast onboarding once offer accepted"),
    ("T9", "Scope creep — additional platforms, users, or systems added mid-engagement",
     "🟡 MED-LOW", "+2 to +6 wks",
     "6-month window blown; extension required",
     "Define scope in writing at kick-off; any additions require contract amendment"),
]

ll_colors = {
    "🔴 HIGH":      (RED_BG,   RED_FG),
    "🟠 MED-HIGH":  (AMBER_BG, AMBER_FG),
    "🟠 MEDIUM":    (AMBER_BG, AMBER_FG),
    "🟡 MED-LOW":   ("FFFDE7", "F57F17"),
}

for idx, (risk, threat, likelihood, weeks, impact, mitigation) in enumerate(threats, 17):
    bg, fg = ll_colors.get(likelihood, (GRAY_BG, MID))
    for ci, val in enumerate([risk, threat, likelihood, weeks, impact, mitigation], 1):
        c = ws.cell(row=idx, column=ci, value=val)
        c.border    = b()
        c.alignment = ctr if ci in (1,3,4) else lft
        if ci == 3:
            c.fill = fill(bg)
            c.font = Font(name="Arial", bold=True, size=10, color=fg)
        elif ci == 1:
            c.fill = fill(RED_BG)
            c.font = Font(name="Arial", bold=True, size=11, color=RED_FG)
        elif ci == 2:
            c.fill = fill(WHITE)
            c.font = Font(name="Arial", bold=True, size=10, color=DARK)
        elif ci == 6:
            c.fill = fill(GREEN_BG)
            c.font = Font(name="Arial", size=10, color=GREEN_FG, italic=True)
        else:
            c.fill = fill(WHITE)
            c.font = Font(name="Arial", size=10, color=MID)
    ws.row_dimensions[idx].height = 42

ws.row_dimensions[26].height = 12

# ── SECTION: REALISTIC TIMELINE SCENARIOS ────────────────────────────
ws.merge_cells("A27:F27")
ws["A27"] = "REALISTIC OUTCOME SCENARIOS"
ws["A27"].font      = Font(name="Arial", bold=True, color=WHITE, size=12)
ws["A27"].fill      = fill(ACCENT2)
ws["A27"].alignment = lft
ws.row_dimensions[27].height = 24

for i, h in enumerate(["Scenario", "Duration", "Likelihood", "What Has to Be True", "Recommendation", ""], 1):
    if i == 6: continue
    c = ws.cell(row=28, column=i, value=h)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill      = fill(ACCENT)
    c.alignment = ctr
    c.border    = b()
ws.row_dimensions[28].height = 28

scenarios = [
    ("Best Case",    "6 months",   "🟡 20%",
     "Access granted fast; dependency map clean; OS version confirmed Week 1; IT responsive; no app compat issues; UAT smooth",
     "Plan for this; don't count on it"),
    ("Most Likely",  "7–8 months", "🔴 55%",
     "1–2 threats materialize (most likely: IT provisioning delay + dependency gaps); managed well but scope needs extension",
     "Build extension language into contract from day one; flag at kick-off"),
    ("Extended",     "9–10 months","🟠 20%",
     "Multiple threats hit: app compat issue + UAT delays + dependency rework; no major disasters but steady friction",
     "Raise at 3-month check-in if pace is behind; don't wait until Week 24"),
    ("At Risk",      "10+ months", "🟡 5%",
     "Major undiscovered dependency or app incompatibility requires platform version upgrades or vendor involvement",
     "Identify this risk in Week 3 Discovery report; surface immediately to Priya"),
]

sc_fills = [GREEN_BG, AMBER_BG, RED_BG, "FCE4EC"]
sc_fgs   = [GREEN_FG, AMBER_FG, RED_FG, "880E4F"]

for idx, (scenario, duration, likelihood, truth, rec) in enumerate(scenarios, 29):
    bg = sc_fills[idx-29]
    fg = sc_fgs[idx-29]
    for ci, val in enumerate([scenario, duration, likelihood, truth, rec, ""], 1):
        if ci == 6: continue
        real_ci = ci
        c = ws.cell(row=idx, column=real_ci, value=val)
        c.border    = b()
        c.alignment = ctr if ci in (1,2,3) else lft
        if ci in (1,2):
            c.fill = fill(bg)
            c.font = Font(name="Arial", bold=True, size=11, color=fg)
        elif ci == 3:
            c.fill = fill(bg)
            c.font = Font(name="Arial", bold=True, size=10, color=fg)
        elif ci == 5:
            c.fill = fill(GREEN_BG)
            c.font = Font(name="Arial", size=10, color=GREEN_FG, italic=True)
        else:
            c.fill = fill(WHITE)
            c.font = Font(name="Arial", size=10, color=MID)
    ws.row_dimensions[idx].height = 48

ws.row_dimensions[33].height = 12

# ── BOTTOM NOTE ───────────────────────────────────────────────────────
ws.merge_cells("A34:F34")
ws["A34"] = (
    "KEY RECOMMENDATION: Build a 2-month extension option into the contract from day one. "
    "The Most Likely scenario (7–8 months) is not a failure — it is a realistic outcome for a "
    "migration of this complexity with a new IT domain, ~140 users, fragmented dependency tracking, "
    "and an OS upgrade as a prerequisite. The 6-month window is achievable only if most of the "
    "HIGH-likelihood threats are resolved in the first 3 weeks."
)
ws["A34"].font      = Font(name="Arial", bold=True, size=10, color=WHITE)
ws["A34"].fill      = fill(ACCENT)
ws["A34"].alignment = lft
ws.row_dimensions[34].height = 52
ws["A34"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb.save("/mnt/user-data/outputs/Northstar Analytics Group_Migration_Timeline_Analysis.xlsx")
print("Done.")
