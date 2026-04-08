from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Migration Checklist"

ACCENT  = "1B4F72"
ACCENT2 = "2471A3"
LT_BLUE = "D6EAF8"
LT_GRAY = "F4F6F7"
GOLD_BG = "FEF9E7"
GOLD    = "B7950B"
WHITE   = "FFFFFF"
DARK    = "1A1A1A"
MID     = "444444"

C_SOLO     = "FFFFFF"
C_PARALLEL = "E8F5E9"
C_DELEGATE = "E3F2FD"
C_JOINT    = "FFF8E1"

S_NOT     = "F5F5F5"
S_WIP     = "FFF9C4"
S_DONE    = "C8E6C9"
S_BLOCKED = "FFCDD2"

def fill(h): return PatternFill("solid", fgColor=h)
def afont(bold=False, color=DARK, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def thin_side(): return Side(style="thin", color="CCCCCC")
def tborder(): return Border(left=thin_side(), right=thin_side(), top=thin_side(), bottom=thin_side())

center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wr = Alignment(horizontal="left",   vertical="center", wrap_text=True)

col_widths = {"A":5,"B":32,"C":48,"D":16,"E":14,"F":14,"G":13,"H":18,"I":26}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Title
ws.merge_cells("A1:I1")
ws["A1"] = "Northstar Analytics Group  —  ANALYTICS PLATFORM SEPARATION MIGRATION CHECKLIST"
ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws["A1"].fill = fill(ACCENT)
ws["A1"].alignment = center
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:I2")
ws["A2"] = "SAS  |  Spotfire  |  Power BI Gateway     Dev → Sandbox → Prod     Windows Server upgrade to current supported version (2022/2025 TBD)"
ws["A2"].font = Font(name="Arial", bold=False, color=WHITE, size=11, italic=True)
ws["A2"].fill = fill(ACCENT2)
ws["A2"].alignment = center
ws.row_dimensions[2].height = 22

ws.merge_cells("A3:I3")
ws["A3"] = (
    "ROW COLOR:   WHITE = You (solo)     GREEN = Parallel (can overlap)     "
    "BLUE = Delegate to Evan/SAS Admin/IT     AMBER = Joint / collaborative          "
    "STATUS:   Not Started  |  In Progress  |  Complete  |  Blocked"
)
ws["A3"].font = Font(name="Arial", size=9, color=DARK)
ws["A3"].fill = fill("F0F4F8")
ws["A3"].alignment = left_wr
ws.row_dimensions[3].height = 28

ws.row_dimensions[4].height = 6

headers = ["#","Task","Completion Criteria / Definition of Done","Phase","Owner","Parallel?","Status","Platform(s)","Notes / Dependencies"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = fill(ACCENT)
    c.alignment = center
    c.border = tborder()
ws.row_dimensions[5].height = 32
ws.freeze_panes = "A6"

dv = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,Blocked"', allow_blank=False)
ws.add_data_validation(dv)

STATUS_FILL = {"Not Started":S_NOT,"In Progress":S_WIP,"Complete":S_DONE,"Blocked":S_BLOCKED}

rows = [
    ("PHASE 0  ·  PRE-START", None),
    (1,"Review all materials from Priya & Evan interviews",
       "Notes consolidated; open questions logged; nothing assumed",
       "0 · Pre-Start","You","No","Not Started","All",
       "Include CMDB fragmentation finding from Evan as a Day 1 risk item",C_SOLO),
    (2,"Confirm contract signed and access provisioned",
       "Written confirmation received; start date locked; NDAs complete",
       "0 · Pre-Start","You + Palmer Group","No","Not Started","All",
       "Kristi / Sarah at The Palmer Group",C_JOINT),
    (3,"Schedule kick-off with Priya and Evan",
       "Meeting on calendar; agenda sent in advance",
       "0 · Pre-Start","You","No","Not Started","All",
       "Do this Day 1",C_SOLO),
    (4,"Obtain environment access: Dev, Sandbox, Prod",
       "Can log in to all three tiers on all three platforms; access verified",
       "0 · Pre-Start","Evan / IT","Yes","Not Started","SAS / Spotfire / PBI",
       "Delegate request to Evan; follow up daily until resolved",C_DELEGATE),
    (5,"Confirm target Windows Server version with Priya and Evan",
       "Written agreement on OS target version (2022, 2025, or other); documented and signed off",
       "0 · Pre-Start","You + Priya + Evan","No","Not Started","Infrastructure",
       "Evan indicated latest supported version may be target — get this confirmed in writing before any planning. Current latest is Server 2025.",C_JOINT),

    ("PHASE 1  ·  DISCOVERY", None),
    (6,"Document current SAS environment (version, host, config, scheduled jobs)",
       "SAS version, grid/Viya config, all scheduled jobs, data sources fully documented",
       "1 · Discovery","You + SAS Admin","Yes","Not Started","SAS",
       "Evan has been primary SAS admin ~2.5 years since predecessor retired — no formal handoff occurred. Gaps are likely.",C_JOINT),
    (7,"Document current Spotfire environment (version, host, data connections, auth)",
       "Spotfire server version, all data connections, user groups, auth method documented",
       "1 · Discovery","Evan","Yes","Not Started","Spotfire",
       "Evan leads; you validate output against what you expect to see",C_DELEGATE),
    (8,"Document current Power BI Gateway (version, data sources, credentials, auth)",
       "Gateway version, all registered data sources, service accounts / AD groups documented",
       "1 · Discovery","Evan","Yes","Not Started","Power BI",
       "Auth setup and credential storage are highest risk — press for specifics",C_DELEGATE),
    (9,"Audit and consolidate dependency tracking across all platforms",
       "Full inventory of where dependency/config tracking currently lives; gaps and overlaps identified; single working map produced",
       "1 · Discovery","You + Evan","No","Not Started","All",
       "IMPORTANT: Evan indicated dependency tracking is fragmented with no single source of truth — 'all over the place.' There is no reliable CMDB baseline to start from. Audit ALL sources (ServiceNow if present, Datadog, spreadsheets, tribal knowledge) and build a unified map before touching anything.",C_JOINT),
    (10,"Trace and document all database connections for the separated environment",
       "Full list of DB endpoints confirmed; which persist post-separation vs. which need new instances identified",
       "1 · Discovery","You + DBA","Yes","Not Started","All",
       "~140 users (35% of 400) moving to new IT domain. Confirm which DB connections follow them vs. stay with parent.",C_JOINT),
    (11,"Inventory all servers in scope for OS upgrade and separation",
       "Hostnames, roles, installed software, current OS patch level documented for every server across all 3 platforms",
       "1 · Discovery","Evan / IT","Yes","Not Started","Infrastructure",
       "Scope: how many servers total? All 3 platforms on same hosts or separate?",C_DELEGATE),
    (12,"Assess all documentation gaps and flag risks to Priya",
       "Gap report delivered; risks ranked; Priya has acknowledged and prioritized",
       "1 · Discovery","You","No","Not Started","All",
       "Fragmented dependency tracking from #9 will likely dominate this report. Do not proceed to Planning until reviewed.",C_SOLO),

    ("PHASE 2  ·  PLANNING", None),
    (13,"Confirm OS upgrade sequence: upgrade first, then separate (as Evan agreed)",
       "Written agreement that OS upgrade precedes separation work; sequencing documented",
       "2 · Planning","You + Priya + Evan","No","Not Started","Infrastructure",
       "Evan agreed: upgrade OS before doing the spinoff separation. Lock this in writing.",C_JOINT),
    (14,"Confirm target OS version and obtain IT approval",
       "Target version (2022/2025/other) formally approved by IT; procurement/licensing confirmed",
       "2 · Planning","Priya + IT","No","Not Started","Infrastructure",
       "Delegate approval to Priya; you need the answer before build work begins",C_DELEGATE),
    (15,"Build consolidated dependency map as single source of truth",
       "All platform dependencies, DB connections, service accounts, and AD groups in one verified document; Evan and Priya sign off",
       "2 · Planning","You","No","Not Started","All",
       "This replaces the fragmented tracking Evan described. Non-negotiable foundation for safe migration.",C_SOLO),
    (16,"Confirm migration sequence across platforms (SAS / Spotfire / PBI order)",
       "Agreed platform migration order documented and signed off by Priya",
       "2 · Planning","You + Priya","No","Not Started","All",
       "Dependencies between platforms may dictate order — validate against the map from #15",C_JOINT),
    (17,"Build rollback plan for each platform and each environment tier",
       "Rollback steps documented for Dev, Sandbox, and Prod for all 3 platforms; reviewed by Evan",
       "2 · Planning","You","Yes","Not Started","All",
       "Script-level diffs captured before any change. No exceptions.",C_SOLO),
    (18,"Establish change control process for migration duration",
       "Change log template in place; agreed tool (Jira or equivalent) configured; Evan and Priya aware",
       "2 · Planning","You","No","Not Started","All",
       "Propose Jira if available; minimum: shared timestamped log",C_SOLO),
    (19,"Establish source control for all scripts and configs being migrated",
       "Git repo (or agreed tool) initialized; baseline snapshots committed before any changes made",
       "2 · Planning","You","No","Not Started","SAS / Spotfire / PBI",
       "Zero changes without a committed baseline.",C_SOLO),
    (20,"Review and validate Priya's migration playbook draft",
       "Playbook reviewed; gaps annotated; feedback returned to Priya",
       "2 · Planning","You + Priya","Yes","Not Started","All",
       "Playbook was in progress per Priya — do not wait for final version",C_JOINT),
    (21,"Prepare and present 2-week next-steps plan to Priya and Evan",
       "Presentation delivered within 2 weeks of start; plan accepted or revised with agreement",
       "2 · Planning","You","No","Not Started","All",
       "Evan specifically requested this deliverable. Include OS upgrade sequencing, separation scope (~140 users / new IT domain), and dependency audit findings.",C_SOLO),
    (22,"Define success criteria and sign-off process for each environment",
       "Written definition of done for Dev, Sandbox, and Prod accepted by Priya",
       "2 · Planning","You + Priya","No","Not Started","All",
       "Protects you at contract end",C_JOINT),

    ("PHASE 3  ·  OS UPGRADE (Pre-Separation)", None),
    (23,"Stand up new Windows Server (target version) for each platform in Dev",
       "Servers provisioned, patched, domain-joined, accessible; confirmed by IT",
       "3 · OS Upgrade","IT / Evan","Yes","Not Started","Infrastructure",
       "Target version TBD — do not begin until #14 is complete",C_DELEGATE),
    (24,"Migrate SAS to upgraded OS in Dev; validate all jobs and configs",
       "SAS running on new OS in Dev; all scheduled jobs execute; output matches baseline",
       "3 · OS Upgrade","You + SAS Admin","No","Not Started","SAS",
       "Validate against documented baseline from Phase 1 #6",C_JOINT),
    (25,"Migrate Spotfire to upgraded OS in Dev; validate connections and auth",
       "Spotfire running on new OS; data connections restored; test users authenticated",
       "3 · OS Upgrade","Evan","Yes","Not Started","Spotfire",
       "Evan leads; you validate",C_DELEGATE),
    (26,"Migrate Power BI Gateway to upgraded OS in Dev; validate all data sources",
       "Gateway registered on new OS; all data sources active; test refresh succeeds",
       "3 · OS Upgrade","You + Evan","Yes","Not Started","Power BI",
       "Re-test all service accounts and auth — do not assume config carries cleanly",C_JOINT),
    (27,"Promote OS upgrade through Sandbox and Prod with full regression each tier",
       "All platforms running on upgraded OS in Sandbox and Prod; full regression passed; Priya sign-off obtained",
       "3 · OS Upgrade","You + Evan + Priya","No","Not Started","All",
       "Complete OS upgrade across all tiers before beginning separation work",C_JOINT),

    ("PHASE 4  ·  DEV SEPARATION BUILD", None),
    (28,"Stand up new separated Dev environment in new IT domain",
       "New domain infrastructure provisioned; accessible; confirmed by IT",
       "4 · Dev Separation","IT / Evan","Yes","Not Started","Infrastructure",
       "New company = new IT domain. Confirm AD, DNS, networking scope with IT before start.",C_DELEGATE),
    (29,"Install and configure SAS in separated Dev on new domain",
       "SAS installed; config matches source baseline; test jobs run successfully",
       "4 · Dev Separation","You + SAS Admin","No","Not Started","SAS",
       "~140 users (35% of 400) in new domain. Validate user access against new AD.",C_JOINT),
    (30,"Install and configure Spotfire in separated Dev on new domain",
       "Spotfire running; data connections restored; user groups migrated to new domain; auth confirmed",
       "4 · Dev Separation","Evan","Yes","Not Started","Spotfire",
       "Evan leads; you validate against Phase 1 documentation",C_DELEGATE),
    (31,"Install and configure Power BI Gateway in separated Dev on new domain",
       "Gateway registered; all data sources reconnected; new domain credentials validated; test refresh succeeds",
       "4 · Dev Separation","You + Evan","Yes","Not Started","Power BI",
       "Auth and service accounts must be recreated in new domain — highest risk item",C_JOINT),
    (32,"Reconnect and validate all DB dependencies in separated Dev (per unified map #15)",
       "Every connection on the consolidated dependency map verified working in separated Dev",
       "4 · Dev Separation","You","No","Not Started","All",
       "This is where fragmented legacy tracking will cause problems. Trust only the map from #15.",C_SOLO),
    (33,"Full regression in separated Dev: jobs, reports, data refreshes",
       "All known scheduled jobs and reports run successfully; output matches baseline",
       "4 · Dev Separation","You","No","Not Started","All",
       "Capture baseline from current prod first for comparison",C_SOLO),
    (34,"Document all issues found in separated Dev; confirm resolutions",
       "Issue log complete; all critical items resolved before Sandbox",
       "4 · Dev Separation","You","No","Not Started","All",
       "Nothing moves to Sandbox with open critical issues",C_SOLO),

    ("PHASE 5  ·  SANDBOX SEPARATION", None),
    (35,"Stand up separated Sandbox environment in new IT domain",
       "Sandbox provisioned and accessible; confirmed by IT",
       "5 · Sandbox","IT / Evan","Yes","Not Started","Infrastructure",
       "Can run in parallel with Dev regression",C_DELEGATE),
    (36,"Promote all platform configs from separated Dev to separated Sandbox",
       "All 3 platforms running in Sandbox; change log updated; diffs captured",
       "5 · Sandbox","You + Evan","No","Not Started","All",
       "Promote only — no manual changes in Sandbox",C_JOINT),
    (37,"Reconnect and validate all DB dependencies in separated Sandbox",
       "Every connection on the consolidated dependency map verified in Sandbox",
       "5 · Sandbox","You","No","Not Started","All",
       "Run the same checklist as Dev #32 — again",C_SOLO),
    (38,"Full UAT in separated Sandbox with ~140 affected users / stakeholders",
       "Key users tested and signed off; no critical issues open",
       "5 · Sandbox","Priya / Users","Yes","Not Started","All",
       "Priya owns sign-off; you support and document results",C_DELEGATE),
    (39,"Sandbox sign-off from Priya before Prod cutover",
       "Written sign-off from Priya confirming Sandbox separation complete",
       "5 · Sandbox","Priya","No","Not Started","All",
       "Do not schedule Prod cutover without this in writing",C_DELEGATE),

    ("PHASE 6  ·  PROD SEPARATION CUTOVER", None),
    (40,"Schedule Prod cutover window; communicate to all ~140 affected users",
       "Cutover date/time confirmed; maintenance window communicated; rollback decision point agreed",
       "6 · Prod Cutover","You + Priya","No","Not Started","All",
       "Pick low-usage window; have rollback steps open before starting",C_JOINT),
    (41,"Take final snapshot / backup of current Prod before cutover",
       "Full backup of all 3 platforms confirmed; integrity verified; stored securely",
       "6 · Prod Cutover","You + IT","No","Not Started","All",
       "This is your rollback safety net — do not skip or abbreviate",C_JOINT),
    (42,"Stand up separated Prod environment in new IT domain",
       "Prod servers provisioned, patched, domain-joined, accessible",
       "6 · Prod Cutover","IT / Evan","No","Not Started","Infrastructure",
       "Should be provisioned before cutover window opens",C_DELEGATE),
    (43,"Promote all platform configs to separated Prod via change control",
       "All 3 platforms running in separated Prod; change log updated",
       "6 · Prod Cutover","You + Evan","No","Not Started","All",
       "Follow exact same sequence as Sandbox — no improvising",C_JOINT),
    (44,"Reconnect and validate all DB dependencies in separated Prod",
       "Every connection on the consolidated dependency map verified live in Prod",
       "6 · Prod Cutover","You","No","Not Started","All",
       "Run the checklist a third time. This is Prod.",C_SOLO),
    (45,"Go / No-Go decision with Priya at rollback decision point",
       "Explicit go/no-go documented; rollback executed within agreed window if no-go",
       "6 · Prod Cutover","You + Priya","No","Not Started","All",
       "Have rollback steps open and ready before this conversation",C_JOINT),

    ("PHASE 7  ·  VALIDATION", None),
    (46,"Full regression in separated Prod: jobs, reports, data refreshes",
       "All scheduled jobs completed; report outputs match pre-migration baseline",
       "7 · Validation","You","No","Not Started","All",
       "Compare against baseline captured before cutover",C_SOLO),
    (47,"Monitor platform health for 48-72 hours post-cutover",
       "No critical alerts; any anomalies investigated and resolved or accepted",
       "7 · Validation","You + Evan","No","Not Started","All",
       "Agree on alert routing with Evan during stabilization window",C_JOINT),
    (48,"Confirm availability and monitoring in place for all separated Prod environments",
       "Monitoring configured; availability baseline established; Priya has visibility",
       "7 · Validation","You","No","Not Started","All",
       "Do not leave Prod without measurement in place",C_SOLO),
    (49,"Collect user feedback from ~140 separated users; resolve reported issues",
       "Feedback loop active; all P1/P2 issues resolved; log closed or transitioned",
       "7 · Validation","You + Priya","No","Not Started","All",
       "Establish feedback channel on Day 1 of Prod — don't wait for problems to find you",C_JOINT),
    (50,"Prod sign-off from Priya",
       "Written sign-off from Priya confirming separated Prod migration complete and stable",
       "7 · Validation","Priya","No","Not Started","All",
       "Get this before contract end or extension conversation",C_DELEGATE),

    ("PHASE 8  ·  CLOSEOUT", None),
    (51,"Finalize consolidated dependency map and hand off as ongoing CMDB baseline",
       "Clean, verified dependency document delivered to Evan and Priya as the new single source of truth",
       "8 · Closeout","You","No","Not Started","All",
       "This directly addresses the fragmented tracking problem identified in Discovery. It is one of the most valuable things you leave behind.",C_SOLO),
    (52,"Finalize and hand off all migration documentation",
       "Complete runbooks, config docs, change logs delivered to Evan and Priya",
       "8 · Closeout","You","No","Not Started","All",
       "Documentation is part of the deliverable, not a nice-to-have",C_SOLO),
    (53,"Confirm decommission plan for old parent-company environments",
       "Decommission timeline agreed; IT has plan; old servers not removed until sign-off",
       "8 · Closeout","Priya + IT","Yes","Not Started","Infrastructure",
       "Delegate — not your call, but flag it explicitly",C_DELEGATE),
    (54,"Contract extension or close-out discussion with Palmer Group",
       "Extension terms confirmed in writing, OR close-out acknowledged and final invoice submitted",
       "8 · Closeout","You + Palmer Group","No","Not Started","All",
       "Kristi / Sarah at The Palmer Group",C_JOINT),
]

row_num = 6
for item in rows:
    if len(item) == 2 and item[1] is None:
        ws.merge_cells(f"A{row_num}:I{row_num}")
        c = ws.cell(row=row_num, column=1, value=item[0])
        c.font = Font(name="Arial", bold=True, color=WHITE, size=12)
        c.fill = fill(ACCENT2)
        c.alignment = left_wr
        c.border = tborder()
        ws.row_dimensions[row_num].height = 22
        row_num += 1
        continue

    num, task, criteria, phase, owner, parallel, status, platforms, notes, row_color = item
    values = [num, task, criteria, phase, owner, parallel, status, platforms, notes]
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=ci, value=val)
        c.fill = fill(row_color)
        c.border = tborder()
        c.alignment = left_wr if ci in (2,3,9) else center
        if ci == 1:
            c.font = Font(name="Arial", bold=True, size=11, color=ACCENT)
        elif ci == 2:
            c.font = Font(name="Arial", bold=True, size=11, color=DARK)
        elif ci == 7:
            c.fill = fill(STATUS_FILL.get(val, S_NOT))
            c.font = Font(name="Arial", bold=True, size=10, color=DARK)
        elif ci == 5:
            c.font = Font(name="Arial", size=10, color="1A237E", italic=True)
        elif ci == 6:
            if val == "Yes":
                c.fill = fill("C8E6C9")
                c.font = Font(name="Arial", bold=True, size=10, color="1B5E20")
            else:
                c.font = Font(name="Arial", size=10, color=MID)
        else:
            c.font = Font(name="Arial", size=10, color=MID)
    dv.sqref = f"G{row_num}"
    ws.row_dimensions[row_num].height = 42
    row_num += 1

# Legend sheet
ws2 = wb.create_sheet("Legend & Color Key")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 55
ws2.merge_cells("A1:B1")
ws2["A1"] = "LEGEND & COLOR KEY"
ws2["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=13)
ws2["A1"].fill = fill(ACCENT)
ws2["A1"].alignment = center
ws2.row_dimensions[1].height = 28

legend = [
    ("ROW COLOR (Execution)", ""),
    ("White — You (solo)",          "You own and execute this task alone, sequentially"),
    ("Green — Parallel",            "Can run concurrently with other work"),
    ("Blue — Delegate",             "Assign to Evan, SAS Admin, or IT; you validate output"),
    ("Amber — Joint",               "Requires active input from both you and Northstar Analytics Group"),
    ("", ""),
    ("STATUS", ""),
    ("Not Started",  "Work has not begun"),
    ("In Progress",  "Actively being worked on"),
    ("Complete",     "All completion criteria met and verified"),
    ("Blocked",      "Cannot proceed — dependency or issue unresolved"),
    ("", ""),
    ("KEY NOTES", ""),
    ("OS Target",               "Confirmed with Evan: latest supported Windows Server version (2022/2025 TBD). Get written confirmation before planning."),
    ("OS Upgrade Sequencing",   "Evan agreed: upgrade OS first, then execute separation. Phase 3 must complete before Phase 4."),
    ("Dependency Tracking",     "Evan indicated tracking is fragmented — no single source of truth. Phase 1 #9 builds the consolidated map that everything else depends on."),
    ("User Scope",              "~140 users (35% of ~400 total) moving to new IT domain / new company in the separation."),
    ("2-Week Deliverable",      "Evan requested a next-steps presentation within 2 weeks of start. See Phase 2 #21."),
]

colors_map = {
    "White — You (solo)": C_SOLO,
    "Green — Parallel":   C_PARALLEL,
    "Blue — Delegate":    C_DELEGATE,
    "Amber — Joint":      C_JOINT,
    "Not Started":        S_NOT,
    "In Progress":        S_WIP,
    "Complete":           S_DONE,
    "Blocked":            S_BLOCKED,
}
section_labels = {"ROW COLOR (Execution)", "STATUS", "KEY NOTES"}

for i, (label, desc) in enumerate(legend, 2):
    ws2.row_dimensions[i].height = 24
    if label in section_labels:
        ws2.merge_cells(f"A{i}:B{i}")
        c = ws2.cell(row=i, column=1, value=label)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
        c.fill = fill(ACCENT2)
        c.alignment = left_wr
        c.border = tborder()
        ws2.cell(row=i, column=2).border = tborder()
    elif label == "":
        pass
    else:
        bg = colors_map.get(label, "FFFFFF")
        for ci, val in enumerate([label, desc], 1):
            c = ws2.cell(row=i, column=ci, value=val)
            c.font = Font(name="Arial", size=10, color=DARK)
            c.fill = fill(bg)
            c.border = tborder()
            c.alignment = left_wr

wb.save("/mnt/user-data/outputs/Northstar Analytics Group_Migration_Checklist.xlsx")
print("Done.")
